import os
import openpyxl

# 识别结果文件夹
RESULT_DIRS = 'C:/Users/a7825/Desktop/工作空间/跑代码/mac0001/out'
# 正解文件夹
CORRECT_DIRS = 'C:/Users/a7825/Desktop/工作空间/跑代码/mac0001/zhengjie'
# 特征值文件夹
EIGENVALUE_DIRS = 'C:/Users/a7825/Desktop/工作空间/跑代码/mac0001/tezheng'

# 特征值合并后的文件
MERGE_CSV = 'C:/Users/a7825/Desktop/工作空间/跑代码/mac0001/mac0001.csv'


class HandelInfo(object):

    def __init__(self):
        # 当前操作正解的excel文件中的工作表
        self.correct_file_wb_sheet = None

        # 操作的数据集
        self.result_set = ' '  # a
        # 当前操作的识别结果文件
        self.result_file_name = ''  # a01
        self.merge_file = open(MERGE_CSV, 'a+')

    # 解析识别结果
    def parse_out(self):
        # 遍历识别结果中的所有文件
        for file_name in os.listdir(RESULT_DIRS):
            # result_file_name 表示当前处理的识别结果文件 a01.out
            self.result_file_name = file_name.split('.')[0]
            # 从文件名中获取a01,a02,a03,a04
            # result_set 表示当前处理的类集
            if self.result_set != file_name[0]:
                self.result_set = file_name[0]
                # correct_file_wb_sheet 表示打开当前类集对应的正解的excel
                self.correct_file_wb_sheet = openpyxl.load_workbook(
                    os.path.join(CORRECT_DIRS, '%s' % self.result_set + '.xlsx')).active
            # 读取识别结果文件
            with open(os.path.join(RESULT_DIRS, file_name), 'rb') as file:
                # 识别出来的日文
                data = file.readlines(1)[0].decode('utf-8').split(':')[1].strip()
                self.comparison(data)

    # 用识别出来的结果取正解的excel文件中做对比
    def comparison(self, data):
        # csv_file： 识别结果对应的特征值文件
        csv_file = os.path.join(EIGENVALUE_DIRS, '%s' % self.result_file_name + '.wav.csv')
        # sheet最大行数
        for row in range(1, self.correct_file_wb_sheet.max_row + 1):
            value = self.correct_file_wb_sheet.cell(row=row, column=1).value
            if value:
                #if self.strQ2B(value) == self.result_file_name.upper():改了之后就不怕空格了
                if self.strQ2B(value).strip() == self.result_file_name.upper().strip():
                    z = ''.join(self.correct_file_wb_sheet.cell(row=row, column=2).value.split())
                    o = ''.join(data.split()).replace('、', '')
                    if z == o:
                        # 识别出来正确的结果
                        print('识别结果正确: %s--->%s' % (o, z))
                        self.handel_csv(csv_file, True)
                    else:
                        print('识别结果错误: %s--->%s' % (o, z))

                        self.handel_csv(csv_file, False)

    def handel_csv(self, csv, v):
        # 正确打1 错误打0
        with open(csv, 'r') as f:
            if v:
                data = f.read().replace("'unknown'", '1')
                print(data)
                self.merge_file.write(data)
            else:
                data = f.read().replace("'unknown'", '0')
                print(data)

                self.merge_file.write(data)

    def strQ2B(self, ustring):
        """把字符串全角转半角"""
        rstring = ""
        for uchar in ustring:
            inside_code = ord(uchar)
            if inside_code == 0x3000:
                inside_code = 0x0020
            else:
                inside_code -= 0xfee0
            if inside_code < 0x0020 or inside_code > 0x7e:  # 转完之后不是半角字符返回原来的字符
                rstring += uchar
            rstring += chr(inside_code)
        return rstring


if __name__ == '__main__':
    a = HandelInfo()
    a.parse_out()
