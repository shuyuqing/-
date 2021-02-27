#機能名ことにフィルター作成(機能id+機能名)
import openpyxl, os,shutil

# path1 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\PT仕様書兼成績書_フォーマット.xlsm"
#機能idと機能名
path2 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\機能名.xlsx"
# path3 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\PT仕様書"
#生成的文件夹要放在哪里
path4 = r"C:\プロジェクト関係\新しいフォルダー(2)\新しいフォルダー"

# workbook_1 = openpyxl.load_workbook(path1)
workbook = openpyxl.load_workbook(path2)
sheet = workbook["Sheet1"]

row_index_1 = 1
filename='PT仕様書兼成績書'

for row_index in range(6,34):

    kinouid=sheet.cell(row=row_index, column=1).value
    kinoumei=sheet.cell(row=row_index, column=2).value
    # saisyufilename=filename+'_'+kinouid+'_'+kinoumei+'.xlsm'
    # path5=os.path.join(path3, saisyufilename)
    # shutil.copy(path1, path5)
    saisyufilename=os.path.join(path4,kinouid+'_'+kinoumei)

    os.mkdir(saisyufilename)
    # os.mkdir(os.path.join(saisyufilename,"01_修正前"))
    # os.mkdir(os.path.join(saisyufilename,"02_修正後"))

    # worksheet = workbook_1.create_sheet(title='sheet1')
    # workbook_1.save(path5)
    # workbook_1.close()

workbook.close()
