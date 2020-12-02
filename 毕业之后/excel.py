import openpyxl, os

path = r"C:\プロジェクト関係\機能一覧_SS成果物マッチング - コピー.xlsx"
excel_path = r"C:\プロジェクト関係\新規作成.xlsx"

workbook = openpyxl.load_workbook(path)
sheet = workbook["Sheet3"]
worksheet = workbook.create_sheet(title='新規作成')
row_index_1 = 1

for column_index in range(3,140):

    worksheet.cell(row=row_index_1, column=1).value = sheet.cell(row=1, column=column_index).value
    flag_hajimete = True
    flag_maru = False

    for row_index_2 in range(1,167):

        if sheet.cell(row=row_index_2, column=column_index).value == "○":

            flag_maru= True

            if flag_hajimete == True:

                worksheet.cell(row=row_index_1, column=2).value = sheet.cell(row=row_index_2, column=2).value
                flag_hajimete = False

            else:

                worksheet.cell(row=row_index_1, column=1).value = sheet.cell(row=1, column=column_index).value
                worksheet.cell(row=row_index_1, column=2).value = sheet.cell(row=row_index_2, column=2).value
            row_index_1 += 1

    if flag_maru == False:
        row_index_1 += 1

workbook.save(excel_path)
workbook.close()
