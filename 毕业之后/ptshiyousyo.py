#PT仕様書フォーマットのcopy(機能名ごと)
import openpyxl, os,shutil

path1 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\PT仕様書兼成績書_フォーマット.xlsm"

path2 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\機能名.xlsx"
path3 = r"C:\プロジェクト関係\PT仕様書作成_機能名ごと\PT仕様書"

# workbook_1 = openpyxl.load_workbook(path1)
workbook = openpyxl.load_workbook(path2)
sheet = workbook["Sheet1"]

row_index_1 = 1
filename='PT仕様書兼成績書'

for row_index in range(1,124):

    kinouid=sheet.cell(row=row_index, column=1).value
    kinoumei=sheet.cell(row=row_index, column=2).value
    saisyufilename=filename+'_'+kinouid+'_'+kinoumei+'.xlsm'
    path5=os.path.join(path3, saisyufilename)
    shutil.copy(path1, path5)

    # worksheet = workbook_1.create_sheet(title='sheet1')
    # workbook_1.save(path5)
    # workbook_1.close()

workbook.close()
